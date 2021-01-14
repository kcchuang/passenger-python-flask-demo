from flask import Flask, render_template
from flask import Flask, flash, request, redirect, url_for
from werkzeug.utils import secure_filename
from flask import send_from_directory
from flask import send_file, Response
from sqlalchemy import create_engine
import pandas as pd
import cx_Oracle
import os
import xlwt
from openpyxl import Workbook, load_workbook

MyApp = Flask(__name__, static_url_path='/static')

def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)

@MyApp.route("/test", methods=['GET', 'POST'])
def test():
    print("kevin: connection test!")
    print("kevin: request.method = ", str(request.method))
    
    if request.method == 'POST':
        print("pingora: POST connection test!")
        my_dsn = cx_Oracle.makedsn("10.208.220.198",1521,sid="nntpc")
        conn = cx_Oracle.connect(user="system", password="manager", dsn=my_dsn)
        cursor = conn.cursor()
        
        try:
            querystring = """
                SELECT basedb.meter.metr_numb
                FROM basedb.meter
                WHERE metr_numb = '17687585104'
            """
            cursor.execute(querystring)
            for metr_numb in cursor:
                print("pingora: metr_numb: ", metr_numb)
    
        except cx_Oracle.DatabaseError as e:
            conn.close()
            return str(e)
            
        conn.close()
        return '''
            <h1>Oracle 連線成功!!</h1>
        '''

    return '''
        <form method='POST'>  
        <h1>Oracle連線測試</h1>
        <input type=submit value=測試 style="font-size:24px">
        </form>
     '''
    
# for Oracle connection test
@MyApp.route("/pingora", methods=['GET', 'POST'])
def pingora():
    print("pingora: connection test!")
    print("pingora: request.method = ", str(request.method))
    
    if request.method == 'POST':
        print("pingora: POST connection test!")
        my_dsn = cx_Oracle.makedsn("10.208.220.198",1521,sid="nntpc")
        conn = cx_Oracle.connect(user="system", password="manager", dsn=my_dsn)
        cursor = conn.cursor()
        
        try:
            querystring = """
                SELECT basedb.meter.metr_numb
                FROM basedb.meter
                WHERE metr_numb = '17687585104'
            """
            cursor.execute(querystring)
            for metr_numb in cursor:
                print("pingora: metr_numb: ", metr_numb)
    
        except cx_Oracle.DatabaseError as e:
            conn.close()
            return str(e)
            
        conn.close()
        return '''
            <h1>Oracle 連線成功!!</h1>
        '''

    return '''
        <form method='POST'>  
        <h1>Oracle連線測試</h1>
        <input type=submit value=測試 style="font-size:24px">
        </form>
     '''
    
# 2019-02-12 Old Meter Query by Meter Number
@MyApp.route("/meter_no", methods=['GET', 'POST'])
def meter_no():
    print("meter: test!")
    print("meter: request.method = ", str(request.method))
    
    if request.method == 'POST':
        print("meter: Begin POST!")
        meternum = request.form['meternum']
		
        my_dsn = cx_Oracle.makedsn("10.208.220.198",1521,sid="nntpc")
        conn = cx_Oracle.connect(user="system", password="manager", dsn=my_dsn)
        cursor = conn.cursor()
        
        try:
            querystring = """select basedb.meter.metr_numb, 
                basedb.meter.bill_cycl, basedb.meter.tran_cord, 
                nvl(basedb.meter.cust_type, '[Null]'), basedb.meter.vaid_date, 
                nvl(basedb.meter.tele_numb, '[Null]'), basedb.meter.entr_date 
                from basedb.meter where basedb.meter.metr_numb
                = '"""
            querystring += meternum+"'"
            querystring = "show tables"
            cursor.execute(querystring)
            datastr = ''
            
            for a in cursor:
                datastr += a+"<br>"
                
            '''
            for datastr in cursor:
                print("datastr: ", datastr[1])
            '''
            print("datastr: ", datastr)
        except cx_Oracle.DatabaseError as e:
            conn.close()
            return str(e)
            
        conn.close()
        return "<h1>meter 電號查詢結果</h1><div>"+datastr+"</div>"
        
    return '''
        <h1>meter 電號查詢</h1>
        <form method='POST'>  
        <h1 style="font-size:24px">請輸入meter:</h1>
        <input name=meternum style="font-size:24px">
        <input type=submit value=搜尋 style="font-size:24px">
        </form>
    '''

# 2019-02-12 New Meter Query by Cust_type
@MyApp.route("/meter", methods=['GET', 'POST'])
def meter():
    #print("meter: test!")
    #print("meter: request.method = ", str(request.method))
    
    if request.method == 'POST':
        custtype = request.form['custtype']
		
        my_dsn = cx_Oracle.makedsn("10.208.220.198",1521,sid="nntpc")
        conn = cx_Oracle.connect(user="system", password="manager", dsn=my_dsn)
        cursor = conn.cursor()

        try:
            querystring = """select basedb.meter.metr_numb, 
                basedb.meter.bill_cycl, basedb.meter.tran_cord, 
                nvl(basedb.meter.cust_type, '[Null]'), basedb.meter.vaid_date, 
                nvl(basedb.meter.tele_numb, '[Null]'), basedb.meter.entr_date 
                from basedb.meter where basedb.meter.cust_type
                = '"""
            querystring += custtype+"'"
            cursor.execute(querystring)
            
            book = xlwt.Workbook(encoding="utf-8")
            sheet1 = book.add_sheet("Sheet 1")
            sheet1.write(0, 0, "用戶電號")
            sheet1.write(0, 1, "計算日")
            sheet1.write(0, 2, "變壓器座標")
            sheet1.write(0, 3, "用戶種類")
            sheet1.write(0, 4, "電話號碼")
            sheet1.write(0, 5, "生效日期")
            sheet1.write(0, 6, "輸入日期")
            row=1
        
            for metr_numb, bill_cycl, tran_cord, cust_type, vaid_date, tele_numb, entr_date in cursor:
                sheet1.write(row, 0, metr_numb)
                sheet1.write(row, 1, bill_cycl)
                sheet1.write(row, 2, tran_cord)
                sheet1.write(row, 3, cust_type)
                sheet1.write(row, 4, tele_numb)
                sheet1.write(row, 5, vaid_date)
                sheet1.write(row, 6, entr_date)
                row = row + 1

            book.save("trial.xls")
       
        except cx_Oracle.DatabaseError as e:
            conn.close()
            return str(e)
            
        conn.close()
        return send_file("trial.xls", as_attachment=True)
        
    return '''
        <h1>meter 用戶種類查詢</h1>
        <form method='POST'>  
        <h1 style="font-size:24px">請選擇用戶種類:</h1>
        <select data-trigger="" name="custtype" style="font-size:24px">
          <option value="H">H</option>
          <option value="R">R</option> 
          <option value="Y">Y</option>
          <option value="X">X</option>
          <option value="Z">Z</option> 
          <option value="L">L</option>
        </select>
        <input type=submit value=搜尋 style="font-size:24px">
        </form>
    '''

if MyApp.config["DEBUG"]:
    @MyApp.after_request
    def after_request(response):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate, public, max-age=0"
        response.headers["Expires"] = 0
        response.headers["Pragma"] = "no-cache"
        return response

# 2019-05-5 old equinment query from hicustomer
@MyApp.route("/hicustomer", methods=['GET', 'POST'])
def hicustomer_query():
    print("hicustomer: Begin POST!")
    ownerfsc = request.form['ownerfsc']
    
    my_dsn = cx_Oracle.makedsn("10.208.220.198",1521,sid="nntpc")
    conn = cx_Oracle.connect(user="system", password="manager", dsn=my_dsn, encoding='utf8', nencoding='utf8')
    cursor = conn.cursor()

    try:
        querystring = """select basedb.hicustomer.ufid, 
            basedb.hicustomer.ownerfsc, basedb.hicustomer.ownerufid, 
            basedb.hicustomer.ownertpclid, basedb.hicustomer.name1, 
            nvl(basedb.hicustomer.group1, '[Null]'), basedb.hicustomer.hicus_capacity,
            basedb.hicustomer.metr_numb, basedb.hicustomer.vlevel,
            basedb.hicustomer.cabtype 
            from basedb.hicustomer where basedb.hicustomer.ownerfsc
            = '"""
        '''querystring = """select * from basedb.hicustomer where basedb.hicustomer.ownerfsc
            = '"""'''
        querystring += ownerfsc+"'"
        cursor.execute(querystring)
        
        book = xlwt.Workbook(encoding="utf-8")
        sheet1 = book.add_sheet("Sheet 1")
        sheet1.write(0, 0, "設備流水號")
        sheet1.write(0, 1, "擁有者的設備種類")
        sheet1.write(0, 2, "擁有者的設備流水號")
        sheet1.write(0, 3, "擁有者的圖號座標")
        sheet1.write(0, 4, "地標名稱")
        sheet1.write(0, 5, "組別")
        sheet1.write(0, 6, "容量")
        sheet1.write(0, 7, "電號")
        sheet1.write(0, 8, "電壓別")
        sheet1.write(0, 9, "主/副饋線別")
        row=1
    
        for f0,f1,f2,f3,f4,f5,f6,f7,f8,f9 in cursor:
            sheet1.write(row, 0, f0)
            sheet1.write(row, 1, f1)
            sheet1.write(row, 2, f2)
            sheet1.write(row, 3, f3)
            sheet1.write(row, 4, f4)
            sheet1.write(row, 5, f5)
            sheet1.write(row, 6, f6)
            sheet1.write(row, 7, f7)
            sheet1.write(row, 8, f8)
            sheet1.write(row, 9, f9)
            row = row + 1

        book.save("設備種類.xls")

    except cx_Oracle.DatabaseError as e:
        conn.close()
        return str(e)
        
    conn.close()
    return send_file("設備種類.xls", as_attachment=True)

@MyApp.route("/", methods=['GET', 'POST'])
def hello():
    if request.method == 'POST':
        if request.form['action'] == '圖號查詢':
            print("圖號查詢: POST!!!!!!!!!!!")
            filename = request.form['file']
            print( "filename: %s" % filename )
            filename = filename.upper()
            subclass1 = request.form.get('choices-single-defaul')
            ext = request.form.get('choices-ext')
            find_path = ("/home/kevin/kevin-share2/%s" % subclass1)
            search_file = ("%s.%s" % (filename, ext))
            print ("class    %s " % str(subclass1))
            print ("file     %s " % search_file)
            found = find(search_file, find_path)
            print ("found    %s " % found)
            if found == None:
                return '''
                    <h1>找不到!!  請回上一頁重新搜尋</h1>
                '''
            found_path = found.split(search_file)
            #filename = secure_filename(file.filename)
            if "PDF" in search_file:
                return send_file(found)
            else:
                return send_file(found, as_attachment=True)
            # preview file conetent
            return send_from_directory(directory=found_path[0], filename=search_file)
            return redirect(url_for('download_link',
                                    filename=found))
        elif request.form['action'] == '用戶種類':
            print("meter: Begin POST!")
            custtype = request.form['custtype']
		    
            my_dsn = cx_Oracle.makedsn("10.208.220.198",1521,sid="nntpc")
            conn = cx_Oracle.connect(user="system", password="manager", dsn=my_dsn)
            cursor = conn.cursor()

            try:
                querystring = """select basedb.meter.metr_numb, 
                    basedb.meter.bill_cycl, basedb.meter.tran_cord, 
                    nvl(basedb.meter.cust_type, '[Null]'), basedb.meter.vaid_date, 
                    nvl(basedb.meter.tele_numb, '[Null]'), basedb.meter.entr_date 
                    from basedb.meter where basedb.meter.cust_type
                    = '"""
                querystring += custtype+"'"
                cursor.execute(querystring)
                
                book = xlwt.Workbook(encoding="utf-8")
                sheet1 = book.add_sheet("Sheet 1")
                sheet1.write(0, 0, "用戶電號")
                sheet1.write(0, 1, "計算日")
                sheet1.write(0, 2, "變壓器座標")
                sheet1.write(0, 3, "用戶種類")
                sheet1.write(0, 4, "電話號碼")
                sheet1.write(0, 5, "生效日期")
                sheet1.write(0, 6, "輸入日期")
                row=1
            
                for metr_numb, bill_cycl, tran_cord, cust_type, vaid_date, tele_numb, entr_date in cursor:
                    sheet1.write(row, 0, metr_numb)
                    sheet1.write(row, 1, bill_cycl)
                    sheet1.write(row, 2, tran_cord)
                    sheet1.write(row, 3, cust_type)
                    sheet1.write(row, 4, tele_numb)
                    sheet1.write(row, 5, vaid_date)
                    sheet1.write(row, 6, entr_date)
                    row = row + 1

                book.save("用戶種類.xls")
       
            except cx_Oracle.DatabaseError as e:
                conn.close()
                return str(e)
                
            conn.close()
            return send_file("用戶種類.xls", as_attachment=True)
        else:
            print("equipment: Begin POST!")
            equip_type= request.form['equip_type']
		    
            my_dsn = cx_Oracle.makedsn("10.208.220.198",1521,sid="nntpc")
            conn = cx_Oracle.connect(user="system", password="manager", dsn=my_dsn, encoding='utf8', nencoding='utf8')
            cursor = conn.cursor()

            try:
                querystring = """select * from basedb."""
                querystring += equip_type#+" where ROWNUM <= 10"
                cursor.execute(querystring)
                book = xlwt.Workbook(encoding="utf-8")
                sheet1 = book.add_sheet("Sheet 1")
                if equip_type == "POLE":
                    wb = Workbook()
                    sheet1 = wb.active 
                    sheet1.cell(1, 1).value = "設備流水號"
                    sheet1.cell(1, 2).value = "設備型式"
                    sheet1.cell(1, 3).value = "圖號座標"
                    sheet1.cell(1, 4).value = "電桿的地址或桿號"
                    sheet1.cell(1, 5).value = "導線代號1"
                    sheet1.cell(1, 6).value = "導線代號2"
                    sheet1.cell(1, 7).value = "導線代號3"
                    sheet1.cell(1, 8).value = "導線代號4"
                    sheet1.cell(1, 9).value = "桿號"
                    sheet1.cell(1, 10).value = "長度"
                    sheet1.cell(1, 11).value = "X座標"
                    sheet1.cell(1, 12).value = "Y座標"
                    sheet1.cell(1, 13).value = "電壓別"
                    sheet1.cell(1, 14).value = "主/副饋線別"
                    sheet1.cell(1, 15).value = "水平支線米數"
                    sheet1.cell(1, 16).value = "經度"
                    sheet1.cell(1, 17).value = "緯度"
                    sheet1.cell(1, 18).value = "同意書編號"
                    sheet1.cell(1, 19).value = "說明"
                    sheet1.cell(1, 20).value = "備註"
                    row=2
                
                    for f0,f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f11,f12,f13,f14,f15,f16,f17,f18,f19 in cursor:
                        sheet1.cell(row, 1).value = f0
                        sheet1.cell(row, 2).value = f1
                        sheet1.cell(row, 3).value = f2
                        sheet1.cell(row, 4).value = f3
                        sheet1.cell(row, 5).value = f4
                        sheet1.cell(row, 6).value = f5
                        sheet1.cell(row, 7).value = f6
                        sheet1.cell(row, 8).value = f7
                        sheet1.cell(row, 9).value = f8
                        sheet1.cell(row, 10).value = f9
                        sheet1.cell(row, 11).value = f10
                        sheet1.cell(row, 12).value = f11
                        sheet1.cell(row, 13).value = f12
                        sheet1.cell(row, 14).value = f13
                        sheet1.cell(row, 15).value = f14
                        sheet1.cell(row, 16).value = f15
                        sheet1.cell(row, 17).value = f16
                        sheet1.cell(row, 18).value = f17
                        sheet1.cell(row, 19).value = f18
                        sheet1.cell(row, 20).value = f19
                        row = row + 1
                    wb.save("407(電桿).xlsx")
                    conn.close()
                    return send_file("407(電桿).xlsx", as_attachment=True)

                elif equip_type == "DSBNROOM":
                    sheet1.write(0, 0, "設備流水號")
                    sheet1.write(0, 1, "設備型式")
                    sheet1.write(0, 2, "圖號座標")
                    sheet1.write(0, 3, "地址")
                    sheet1.write(0, 4, "建物名稱")
                    sheet1.write(0, 5, "主要饋線")
                    sheet1.write(0, 6, "備用饋線1")
                    sheet1.write(0, 7, "備用饋線2")
                    sheet1.write(0, 8, "門牌號碼")
                    sheet1.write(0, 9, "連絡電話")
                    sheet1.write(0, 10, "X座標")
                    sheet1.write(0, 11, "Y座標")
                    sheet1.write(0, 12, "電壓別")
                    sheet1.write(0, 13, "主/副饋線別")
                    sheet1.write(0, 14, "水平支線米數")
                    sheet1.write(0, 15, "經度")
                    sheet1.write(0, 16, "緯度")
                    row=1
                
                    for f0,f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f11,f12,f13,f14,f15,f16 in cursor:
                        sheet1.write(row, 0, f0)
                        sheet1.write(row, 1, f1)
                        sheet1.write(row, 2, f2)
                        sheet1.write(row, 3, f3)
                        sheet1.write(row, 4, f4)
                        sheet1.write(row, 5, f5)
                        sheet1.write(row, 6, f6)
                        sheet1.write(row, 7, f7)
                        sheet1.write(row, 8, f8)
                        sheet1.write(row, 9, f9)
                        sheet1.write(row, 10, f10)
                        sheet1.write(row, 11, f11)
                        sheet1.write(row, 12, f12)
                        sheet1.write(row, 13, f13)
                        sheet1.write(row, 14, f14)
                        sheet1.write(row, 15, f15)
                        sheet1.write(row, 16, f16)
                        row = row + 1
                    book.save("411(配電室).xls")
                    conn.close()
                    return send_file("411(配電室).xls", as_attachment=True)

                elif equip_type == "LV_TRF_LIGHT":
                    sheet1.write(0, 0, "設備流水號")
                    sheet1.write(0, 1, "供電變壓器組別")
                    sheet1.write(0, 2, "地下/架空別")
                    sheet1.write(0, 3, "圖號座標")
                    sheet1.write(0, 4, "供電變壓器座標")
                    sheet1.write(0, 5, "門牌號碼")
                    sheet1.write(0, 6, "裝置日期")
                    row=1
                
                    for f0,f1,f2,f3,f4,f5,f6 in cursor:
                        sheet1.write(row, 0, f0)
                        sheet1.write(row, 1, f1)
                        sheet1.write(row, 2, f2)
                        sheet1.write(row, 3, f3)
                        sheet1.write(row, 4, f4)
                        sheet1.write(row, 5, f5)
                        sheet1.write(row, 6, f6)
                        row = row + 1
                    book.save("209(交通號誌).xls")
                    conn.close()
                    return send_file("209(交通號誌).xls", as_attachment=True)

                elif equip_type == "SL_STREET_LIGHT":
                    sheet1.write(0, 0, "設備流水號")
                    sheet1.write(0, 1, "圖號座標")
                    sheet1.write(0, 2, "受理號碼")
                    sheet1.write(0, 3, "台電編號")
                    sheet1.write(0, 4, "電號")
                    sheet1.write(0, 5, "種類")
                    sheet1.write(0, 6, "容量")
                    sheet1.write(0, 7, "盞數")
                    sheet1.write(0, 8, "裝設別")
                    sheet1.write(0, 9, "用電別")
                    sheet1.write(0, 10, "燈別")
                    sheet1.write(0, 11, "裝設方式")
                    sheet1.write(0, 12, "用戶名稱")
                    sheet1.write(0, 13, "行政區編號")
                    sheet1.write(0, 14, "收費單位")
                    sheet1.write(0, 15, "服務所")
                    sheet1.write(0, 16, "街道名稱")
                    sheet1.write(0, 17, "申請日期")
                    sheet1.write(0, 18, "送電日期")
                    sheet1.write(0, 19, "電話及備註欄")
                    sheet1.write(0, 20, "普查日期")
                    sheet1.write(0, 21, "廢止日期")
                    sheet1.write(0, 22, "檢驗員")
                    sheet1.write(0, 23, "註記")
                    sheet1.write(0, 24, "擁有者的設備代號")
                    sheet1.write(0, 25, "擁有者的設備流水號")
                    row=1
                
                    for f0,f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f11,f12,f13,f14,f15,f16,f17,f18,f19,f20,f21,f22,f23,f24,f25 in cursor:
                        sheet1.write(row, 0, f0)
                        sheet1.write(row, 1, f1)
                        sheet1.write(row, 2, f2)
                        sheet1.write(row, 3, f3)
                        sheet1.write(row, 4, f4)
                        sheet1.write(row, 5, f5)
                        sheet1.write(row, 6, f6)
                        sheet1.write(row, 7, f7)
                        sheet1.write(row, 8, f8)
                        sheet1.write(row, 9, f9)
                        sheet1.write(row, 10, f10)
                        sheet1.write(row, 11, f11)
                        sheet1.write(row, 12, f12)
                        sheet1.write(row, 13, f13)
                        sheet1.write(row, 14, f14)
                        sheet1.write(row, 15, f15)
                        sheet1.write(row, 16, f16)
                        sheet1.write(row, 17, f7)
                        sheet1.write(row, 18, f8)
                        sheet1.write(row, 19, f9)
                        sheet1.write(row, 20, f10)
                        sheet1.write(row, 21, f11)
                        sheet1.write(row, 22, f12)
                        sheet1.write(row, 23, f13)
                        sheet1.write(row, 24, f14)
                        sheet1.write(row, 25, f15)
                        row = row + 1
                    book.save("311(路燈台帳).xls")
                    conn.close()
                    return send_file("311(路燈台帳).xls", as_attachment=True)
       
                elif equip_type == "SL_STREET_TRAFFIC_ACCOUNT":
                    sheet1.write(0, 0, "設備流水號")
                    sheet1.write(0, 1, "圖號座標")
                    sheet1.write(0, 2, "擁有者的設備代號")
                    sheet1.write(0, 3, "擁有者的設備流水號")
                    sheet1.write(0, 4, "電號")
                    sheet1.write(0, 5, "用戶名稱")
                    sheet1.write(0, 6, "用電別")
                    sheet1.write(0, 7, "最大同時亮燈實測入力數")
                    sheet1.write(0, 8, "裝設方式")
                    sheet1.write(0, 9, "責任分界點地址")
                    sheet1.write(0, 10, "行政區")
                    sheet1.write(0, 11, "普查部門")
                    sheet1.write(0, 12, "普查日期")
                    sheet1.write(0, 13, "註記")
                    row=1
                
                    for f0,f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f11,f12,f13 in cursor:
                        sheet1.write(row, 0, f0)
                        sheet1.write(row, 1, f1)
                        sheet1.write(row, 2, f2)
                        sheet1.write(row, 3, f3)
                        sheet1.write(row, 4, f4)
                        sheet1.write(row, 5, f5)
                        sheet1.write(row, 6, f6)
                        sheet1.write(row, 7, f7)
                        sheet1.write(row, 8, f8)
                        sheet1.write(row, 9, f9)
                        sheet1.write(row, 10, f10)
                        sheet1.write(row, 11, f11)
                        sheet1.write(row, 12, f12)
                        sheet1.write(row, 13, f13)
                        row = row + 1
                    book.save("310(交通燈).xls")
                    conn.close()
                    return send_file("310(交通燈).xls", as_attachment=True)

                elif equip_type == "CV_HOLE":
                    sheet1.write(0, 0, "設備流水號")
                    sheet1.write(0, 1, "型式")
                    sheet1.write(0, 2, "人手孔別")
                    sheet1.write(0, 3, "圖號座標")
                    sheet1.write(0, 4, "地址")
                    sheet1.write(0, 5, "裝置日期")
                    sheet1.write(0, 6, "管轄單位")
                    sheet1.write(0, 7, "座標註記")
                    sheet1.write(0, 8, "孔蓋形式")
                    sheet1.write(0, 9, "設計號碼")
                    sheet1.write(0, 10, "施工號碼")
                    sheet1.write(0, 11, "孔蓋種類")
                    sheet1.write(0, 12, "孔蓋長度(直徑)")
                    sheet1.write(0, 13, "孔蓋寬度")
                    sheet1.write(0, 14, "高低壓共架")
                    sheet1.write(0, 15, "97座標")
                    sheet1.write(0, 16, "經度")
                    sheet1.write(0, 17, "緯度")
                    sheet1.write(0, 18, "設備狀態")
                    sheet1.write(0, 19, "道路屬性")
                    sheet1.write(0, 20, "孔底高")
                    sheet1.write(0, 21, "底盤高")
                    row=1
                
                    for f0,f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f11,f12,f13,f14,f15,f16,f17,f18,f19,f20,f21 in cursor:
                        sheet1.write(row, 0, f0)
                        sheet1.write(row, 1, f1)
                        sheet1.write(row, 2, f2)
                        sheet1.write(row, 3, f3)
                        sheet1.write(row, 4, f4)
                        sheet1.write(row, 5, f5)
                        sheet1.write(row, 6, f6)
                        sheet1.write(row, 7, f7)
                        sheet1.write(row, 8, f8)
                        sheet1.write(row, 9, f9)
                        sheet1.write(row, 10, f10)
                        sheet1.write(row, 11, f11)
                        sheet1.write(row, 12, f12)
                        sheet1.write(row, 13, f13)
                        sheet1.write(row, 14, f14)
                        sheet1.write(row, 15, f15)
                        sheet1.write(row, 16, f16)
                        sheet1.write(row, 17, f7)
                        sheet1.write(row, 18, f8)
                        sheet1.write(row, 19, f9)
                        sheet1.write(row, 20, f10)
                        sheet1.write(row, 21, f11)
                        row = row + 1
                    book.save("403(人手孔).xls")
                    conn.close()
                    return send_file("403(人手孔).xls", as_attachment=True)

            except cx_Oracle.DatabaseError as e:
                conn.close()
                return str(e)
                

    return '''
<!doctype html>
  <head>
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta http-equiv="X-UA-Compatible" content="IE=edge" />
    <meta name="author" content="colorlib.com">

    <style>
      body, html {
        height: 100%;
        marine: 0;
      }
      .bg1 {
        position: relative;
        opacity: 0.65;
        background-position: center;
        background-repeat: no-repeat;
        background-image: url("/static/images/morning.jpg");
        background-size: cover;
      }
    </style>

  </head>
  <title>台電南投圖層搜尋系統</title>
  <body>
    <div class="bg1">
      <img src="/static/images/title.jpg" width="20%" height="20%" align=”right”></img>
      <form id="package_form" action="" method="post">
        <h1>===========圖層查詢===========</h1>
        <div class="input-field first-wrap">
          <div class="input-select">
            <h1 style="font-size:24px">請選擇圖層:</h1>
            <select data-trigger="" name="choices-single-defaul" style="font-size:24px">
              <option>CMMS地下低壓圖</option>
              <option>CMMS地下高壓圖</option>
              <option>CMMS架空圖</option>
              <option>CMMS管路圖</option>
              <option>CMMS路燈台帳圖</option>
              <option>CMMS路燈圖</option>
              <option>地形圖</option>
              <option>人手孔卡</option>
              <option>光纜</option>
              <option>配電室卡</option>
            </select>
            <h1 style="font-size:24px">請選擇副檔名:</h1>
            <select data-trigger="" name="choices-ext" style="font-size:24px">
              <option>PDF</option>
              <option>TIF</option>
            </select>
          </div>
        </div>
        <h1 style="font-size:24px">請輸入圖號:</h1>
        <input name=file style="font-size:24px">
        <input type ="submit" name="action" value="圖號查詢" style="font-size:24px">

        <h1>===========meter 用戶種類查詢===========</h1>
        <h1 style="font-size:24px">請選擇用戶種類:</h1>
        <select data-trigger="" name="custtype" style="font-size:24px">
          <option value="H">H(高壓用戶)</option>
          <option value="R">R(分散式能源)</option> 
          <option value="Y">Y(重用戶)</option>
          <option value="X">X(重要用戶)</option>
          <option value="Z">Z(使用維生器材)</option> 
          <option value="L">L(低壓重點用戶)</option>
        </select>
        <input type ="submit" name="action" value="用戶種類" style="font-size:24px">

        <h1>===========設備查詢===========</h1>
        <h1 style="font-size:24px">請選擇設備種類(資料庫擷取需時間):</h1>
        <select data-trigger="" name="equip_type" style="font-size:24px">
          <option value="POLE">407(電桿)</option>
          <option value="DSBNROOM">411(配電室)</option> 
          <option value="LV_TRF_LIGHT">209(交通號誌)</option> 
          <option value="SL_STREET_LIGHT">311(路燈台帳)</option> 
          <option value="SL_STREET_TRAFFIC_ACCOUNT">310(交通燈)</option> 
          <option value="CV_HOLE">403(人手孔)</option> 
        </select>
        <input type ="submit" name="action" value="設備查詢" style="font-size:24px">
      </form>
    </bg>
    <h1></h1>
    <a href="https://chrome.google.com/webstore/detail/pdf-viewer/oemmndcbldboiebfnladdacbdfmadadm?hl=zh-TW" target="" title="chrome pdf viewer plugin" style="font-size:18px">如果使用chrome點此安裝pdf viewer</a>
    <h1></h1>
    <a href="https://chrome.google.com/webstore/detail/tiff-viewer/fciggfkkblggmebjbekbebbcffeacknj" target="" title="chrome tiff viewer plugin" style="font-size:18px">如果使用chrome點此安裝tiff viewer</a>
  </body>
</html>
    '''

if __name__ == "__main__":
	MyApp.run(debug=True)
